import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('MASTER AIR AGENT RATE 01ST July.2025.xlsx')

print('Available Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"="*100}')
    print(f'SHEET: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)')
    print(f'{"="*100}')
    
    for row_idx in range(1, min(30, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(30, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                col_letter = get_column_letter(col_idx)
                row_data.append(f'{col_letter}:{val}')
        if row_data:
            print(f'R{row_idx}: {" | ".join(row_data)}')
