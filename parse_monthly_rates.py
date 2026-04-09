import argparse
import json
import re
from pathlib import Path

import openpyxl


PROVIDER_NORMALIZATION = {
    'S-DHL': 'Singapore-DHL',
    'S-UPS': 'Singapore-UPS',
    'D-DHL': 'DUBAI-DHL',
    'D-UPS': 'DUBAI-UPS',
    'S-FedEx': 'Singapore-FedEx',
    'D-FedEx': 'DUBAI-FedEx',
    'Master': 'Master',
    'MASTER': 'Master',
    'Master-SF': 'Master-SF',
    'Master -SF': 'Master-SF',
    'Master- SF': 'Master-SF',
    'Master-Nice': 'Master-Nice',
    'D-DHL/Risk': 'DUBAI-DHL/Risk',
}


def read_float(value):
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_text(value):
    if value is None:
        return ''
    return str(value).strip()


def extract_country_code(country_text):
    text = read_text(country_text)
    match = re.search(r'\(([A-Z]{2})\)?', text)
    if match:
        return match.group(1)
    return None


def extract_header_weight(header_text):
    text = read_text(header_text).lower()
    if not text or 'kg' not in text:
        return None
    if 'per' in text:
        return None

    match = re.search(r'(\d+(?:\.\d+)?)', text)
    if not match:
        return None

    return read_float(match.group(1))


def extract_zone_number(zone_text):
    text = read_text(zone_text)
    if not text:
        return None

    match = re.search(r'(\d+)', text)
    if not match:
        return None

    return int(match.group(1))


def detect_rate_zone_columns(ws, header_row):
    zone_columns = []

    for col_idx in range(1, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        normalized = header_text.lower().replace(' ', '')
        if normalized.startswith('zone'):
            zone_number = extract_zone_number(header_text)
            if zone_number is not None:
                zone_columns.append((zone_number, col_idx))

    return zone_columns


def parse_dhl_zone_list(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    country_headers = []
    zone_headers = []
    header_row = 2

    best_header_row = None
    best_country_count = -1
    scan_header_end = min(ws.max_row, 12)

    for candidate_row in range(1, scan_header_end + 1):
        candidate_country_count = 0
        candidate_zone_count = 0

        for col_idx in range(1, ws.max_column + 1):
            header_text = read_text(ws.cell(candidate_row, col_idx).value).lower()
            if not header_text:
                continue

            if 'countries' in header_text:
                candidate_country_count += 1
            elif header_text.startswith('zone'):
                candidate_zone_count += 1

        if candidate_country_count > 0 and candidate_zone_count > 0 and candidate_country_count > best_country_count:
            best_country_count = candidate_country_count
            best_header_row = candidate_row

    if best_header_row is not None:
        header_row = best_header_row

    for col_idx in range(1, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        normalized = header_text.lower()
        if 'countries' in normalized:
            country_headers.append(col_idx)
        elif normalized.startswith('zone'):
            zone_headers.append(col_idx)

    header_pairs = []
    def pick_zone_value_col(zone_header_col):
        candidates = [zone_header_col]
        if zone_header_col + 1 <= ws.max_column:
            candidates.append(zone_header_col + 1)
        if zone_header_col - 1 >= 1:
            candidates.append(zone_header_col - 1)

        best_col = zone_header_col
        best_score = -1
        scan_end = min(ws.max_row, 80)

        for candidate_col in candidates:
            numeric_count = 0
            for row_idx in range(5, scan_end + 1):
                if read_float(ws.cell(row_idx, candidate_col).value) is not None:
                    numeric_count += 1

            if numeric_count > best_score:
                best_score = numeric_count
                best_col = candidate_col

        return best_col

    for index, country_col in enumerate(country_headers):
        next_country_col = country_headers[index + 1] if index + 1 < len(country_headers) else ws.max_column + 1
        candidate_zones = [col for col in zone_headers if country_col < col < next_country_col]
        if not candidate_zones:
            candidate_zones = [col for col in zone_headers if col > country_col]
        zone_col = candidate_zones[0] if candidate_zones else None
        if zone_col is not None:
            header_pairs.append((country_col, pick_zone_value_col(zone_col)))

    zones = []
    for row_idx in range(3, ws.max_row + 1):
        for country_col, zone_col in header_pairs:
            country_text = read_text(ws.cell(row_idx, country_col).value)
            zone_value = read_float(ws.cell(row_idx, zone_col).value)

            if not country_text or zone_value is None:
                continue

            country_code = extract_country_code(country_text)
            if country_code is None:
                continue

            country_name = re.sub(r'\s*\(.*$', '', country_text).strip()
            zones.append({
                'country_code': country_code,
                'country_name': country_name,
                'zone': int(zone_value),
            })

    return {
        'zones': zones,
        'total_zones': len(zones),
    }


def parse_dhl_rates(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    document_zone_columns = detect_rate_zone_columns(ws, 4)
    non_document_zone_columns = detect_rate_zone_columns(ws, 10)

    if not document_zone_columns:
        document_zone_columns = detect_rate_zone_columns(ws, 5)

    if not non_document_zone_columns:
        non_document_zone_columns = document_zone_columns

    document_rows = []
    non_document_rows = []

    for row_idx in range(5, 9):
        weight = read_float(ws.cell(row_idx, 1).value)
        if weight is None:
            continue
        zones = {}
        for zone_number, col_idx in document_zone_columns:
            zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)
        document_rows.append({'weight': weight, 'zones': zones})

    for row_idx in range(11, 71):
        weight = read_float(ws.cell(row_idx, 1).value)
        if weight is None:
            continue
        zones = {}
        for zone_number, col_idx in non_document_zone_columns:
            zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)
        non_document_rows.append({'weight': weight, 'zones': zones})

    per_kg_bands = []
    for row_idx in range(74, 76):
        from_weight = read_float(ws.cell(row_idx, 1).value)
        to_weight = read_float(ws.cell(row_idx, 2).value)
        if from_weight is None or to_weight is None:
            continue

        zones = {}
        for zone_number, col_idx in non_document_zone_columns:
            zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)

        per_kg_bands.append({
            'from': from_weight,
            'to': to_weight,
            'zones': zones,
        })

    return {
        'document': document_rows,
        'non_document': non_document_rows,
        'per_kg_bands': per_kg_bands,
    }


def parse_master(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    data = {
        'providers': {},
        'total_rates': 0,
    }

    header_row = 7
    expected_parcel_weights = [0.5 + (i * 0.5) for i in range(20)]

    parcel_col_by_weight = {}
    per_0_5_col = None
    per_21_col = None
    per_31_col = None

    for col_idx in range(6, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        lower = header_text.lower()

        if 'per' in lower and '0.5' in lower:
            per_0_5_col = col_idx
            continue

        if '21' in lower and 'kg+' in lower:
            per_21_col = col_idx
            continue

        if '31' in lower and 'kg+' in lower:
            per_31_col = col_idx
            continue

        weight = extract_header_weight(header_text)
        if weight is None:
            continue

        normalized_weight = round(weight, 1)
        if normalized_weight in expected_parcel_weights and normalized_weight not in parcel_col_by_weight:
            parcel_col_by_weight[normalized_weight] = col_idx

    current_country = None
    for row_idx in range(9, ws.max_row + 1):
        destination = ws.cell(row_idx, 1).value
        provider_raw = ws.cell(row_idx, 2).value

        if destination and str(destination).strip():
            current_country = str(destination).strip()

        if provider_raw is None or current_country is None:
            continue

        provider_text = str(provider_raw).strip()
        if provider_text in ('Service', 'Provider'):
            continue

        provider_full = PROVIDER_NORMALIZATION.get(provider_text)
        if provider_full is None:
            continue

        doc_0_5 = read_float(ws.cell(row_idx, 3).value)
        doc_1_0 = read_float(ws.cell(row_idx, 4).value)
        doc_add_0_5 = read_float(ws.cell(row_idx, 5).value)

        parcel_rates = {}
        for i in range(20):
            weight_slab = 0.5 + (i * 0.5)
            col_idx = parcel_col_by_weight.get(round(weight_slab, 1), 6 + i)
            parcel_rates[str(weight_slab)] = read_float(ws.cell(row_idx, col_idx).value)

        per_0_5_kg = read_float(ws.cell(row_idx, per_0_5_col or 26).value)
        per_21_kg = read_float(ws.cell(row_idx, per_21_col or 27).value)
        per_31_kg = read_float(ws.cell(row_idx, per_31_col or 28).value)

        if provider_full not in data['providers']:
            data['providers'][provider_full] = {'countries': {}}

        data['providers'][provider_full]['countries'][current_country] = {
            'document': {
                '0.5': doc_0_5,
                '1.0': doc_1_0,
                'add_0.5': doc_add_0_5,
            },
            'parcel': parcel_rates,
            'per_0_5_kg': per_0_5_kg,
            'per_21_kg': per_21_kg,
            'per_31_kg': per_31_kg,
        }
        data['total_rates'] += 1

    return data


def main():
    parser = argparse.ArgumentParser(description='Parse DHL + Master monthly rate sheets to JSON')
    parser.add_argument('--dhl', help='Path to DHL rate Excel file')
    parser.add_argument('--dhl-rate', help='Path to DHL rate Excel file')
    parser.add_argument('--dhl-zone', help='Path to DHL zone Excel file')
    parser.add_argument('--master', required=True, help='Path to Master Excel file')
    parser.add_argument('--output', required=True, help='Output JSON path')
    args = parser.parse_args()

    dhl_rate_path = Path(args.dhl_rate or args.dhl) if (args.dhl_rate or args.dhl) else None
    dhl_zone_path = Path(args.dhl_zone) if args.dhl_zone else None
    master_path = Path(args.master)
    output_path = Path(args.output)

    if dhl_rate_path is None:
        raise FileNotFoundError('DHL rate file not provided.')

    if not dhl_rate_path.exists():
        raise FileNotFoundError(f'DHL rate file not found: {dhl_rate_path}')

    if dhl_zone_path is not None and not dhl_zone_path.exists():
        raise FileNotFoundError(f'DHL zone file not found: {dhl_zone_path}')

    if not master_path.exists():
        raise FileNotFoundError(f'Master file not found: {master_path}')

    parsed = {
        'dhl': parse_dhl_rates(str(dhl_rate_path)),
        'dhl_zones': parse_dhl_zone_list(str(dhl_zone_path)) if dhl_zone_path is not None else {'zones': [], 'total_zones': 0},
        'master': parse_master(str(master_path)),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(parsed, indent=2), encoding='utf-8')

    print('Parsing complete')
    print(f"DHL document slabs: {len(parsed['dhl']['document'])}")
    print(f"DHL non-document slabs: {len(parsed['dhl']['non_document'])}")
    print(f"DHL zone mappings: {parsed['dhl_zones']['total_zones']}")
    print(f"Master entries: {parsed['master']['total_rates']}")


if __name__ == '__main__':
    main()
