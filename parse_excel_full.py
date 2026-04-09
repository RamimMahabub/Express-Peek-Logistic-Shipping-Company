import openpyxl
import json
from fractions import Fraction

wb = openpyxl.load_workbook('MASTER AIR AGENT RATE 01ST July.2025.xlsx')
ws = wb['Table 1']

# Map provider names and normalize variations
PROVIDER_NORMALIZATION = {
    'S-DHL': 'Singapore-DHL',
    'S-UPS': 'Singapore-UPS',
    'D-DHL': 'DUBAI-DHL',
    'D-UPS': 'DUBAI-UPS',
    'S-FedEx': 'Singapore-FedEx',
    'D-FedEx': 'DUBAI-FedEx',
    'Master': 'Master',
    'MASTER': 'Master',
    'Master-SF': 'Master-SF',
    'Master -SF': 'Master-SF',
    'Master- SF': 'Master-SF',
    'Master-Nice': 'Master-Nice',
    'D-DHL/Risk': 'DUBAI-DHL/Risk',
}

# Column indices (1-indexed for openpyxl)
COL_DESTINATION = 1   # A
COL_PROVIDER = 2      # B
COL_DOC_0_5 = 3       # C: Document 0.5 Kg
COL_DOC_1_0 = 4       # D: Document 1 Kg
COL_DOC_ADD_0_5 = 5   # E: Add 0.5 Kg (increment for documents above 1kg)
COL_PARCEL_0_5 = 6    # F: Parcel 0.5 Kg
# F-Y are parcel weight slabs (0.5kg to 10kg): columns 6-25
# Z = 26: Per 0.5 Kg
# AA = 27: Per Kg 21+
# AB = 28: Per Kg 31+

# Read all data
rates_data = {
    'providers': {},
    'by_country': {},
    'total_rates': 0,
}

current_country = None
for row_idx in range(9, ws.max_row + 1):  # Data starts from row 9
    dest_cell = ws.cell(row_idx, COL_DESTINATION).value
    provider_cell = ws.cell(row_idx, COL_PROVIDER).value
    
    # Update current country if it changes
    if dest_cell:
        current_country = str(dest_cell).strip()
    
    if not provider_cell or not current_country:
        continue
    
    provider = provider_cell.strip()
    
    # Skip invalid providers
    if provider not in ['Service', 'Provider'] and not PROVIDER_NORMALIZATION.get(provider):
        continue
    if provider in ['Service', 'Provider']:
        continue
    
    provider_full = PROVIDER_NORMALIZATION[provider]
    
    # Read document rates (columns C, D, E)
    doc_0_5 = ws.cell(row_idx, COL_DOC_0_5).value
    doc_1_0 = ws.cell(row_idx, COL_DOC_1_0).value
    doc_add_0_5 = ws.cell(row_idx, COL_DOC_ADD_0_5).value
    
    # Read parcel rates (columns F-Y for 0.5-10kg)
    parcel_rates = {}
    for i in range(20):  # 0.5 to 10kg in 0.5kg increments = 20 values (F through Y)
        col_idx = COL_PARCEL_0_5 + i
        val = ws.cell(row_idx, col_idx).value
        weight_slab = 0.5 + (i * 0.5)
        parcel_rates[weight_slab] = val
    
    # Read per-0.5kg, per-21kg+, per-31kg+ (columns Z, AA, AB)
    per_0_5_kg = ws.cell(row_idx, 26).value   # Z
    per_21_kg = ws.cell(row_idx, 27).value    # AA
    per_31_kg = ws.cell(row_idx, 28).value    # AB
    
    # Store data
    if provider_full not in rates_data['providers']:
        rates_data['providers'][provider_full] = {
            'countries': {}
        }
    
    country_data = {
        'document': {
            '0.5': doc_0_5,
            '1.0': doc_1_0,
            'add_0.5': doc_add_0_5,
        },
        'parcel': parcel_rates,
        'per_0_5_kg': per_0_5_kg,
        'per_21_kg': per_21_kg,
        'per_31_kg': per_31_kg,
    }
    
    rates_data['providers'][provider_full]['countries'][current_country] = country_data
    rates_data['total_rates'] += 1

# Print summary
print(f"Parsed {rates_data['total_rates']} rate entries")
print(f"\nProviders found: {list(rates_data['providers'].keys())}")

# Save to JSON for import
with open('excel_rates_parsed.json', 'w') as f:
    json.dump(rates_data, f, indent=2)

print("\nSaved to excel_rates_parsed.json")

# Show sample
print("\nSample (Australia S-DHL):")
if 'Master' in rates_data['providers']:
    if 'Australia' in rates_data['providers']['Master']['countries']:
        au_data = rates_data['providers']['Master']['countries']['Australia']
        print(json.dumps(au_data, indent=2))
